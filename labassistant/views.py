from django.shortcuts import render,HttpResponse
from faculty.models import TimeSlot,Lecture,Subject,Year,Division,DaysOfWeek,Batch,Room
from django.db.models import DurationField, F, ExpressionWrapper
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
import json
from django.core import serializers
from django.http import JsonResponse, QueryDict
from django.conf import settings

from PyPDF2 import PdfFileWriter, PdfFileReader
from collections import OrderedDict
import io
import os
from datetime import datetime, timedelta
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import cloudconvert

@login_required
def get_timetable(request):
	if request.user.is_staff:
		errors = None
		if(request.method == 'POST'):
			day_names = list(DaysOfWeek.objects.values_list('day_name', flat=True))[1:-1]
			form_data = OrderedDict()

			for name in day_names:
				stat_check = request.POST.get('form[' + name + ']', False)
				if stat_check:
					form_data[name] = QueryDict(request.POST.get('form[' + name + ']', False).encode('ASCII'))
				else:
					form_data[name] = QueryDict()

			exclude_list = ['csrfmiddlewaretoken', 'termopt', 'yearopt', 'dayopt', 'divopt']
			year = ""
			div = ""
			term = ""

			for key, value in form_data.items():

				day_name = key
				qdict = value
				if len(qdict) > 0:
					for data_key, data_value in dict(qdict).items():

						if data_key not in exclude_list:
							key_list = list(map(lambda x: x.strip(), data_key.split('-')))

							lecture, room_id = data_value
							if lecture != '0' and room_id != 'None': # these are the values to be plotted on pdf
								sub_id, fac_id = lecture.split("-")
								s = Subject.objects.get(pk=int(sub_id))
								u = User.objects.get(pk=int(fac_id))
								r = Room.objects.get(pk=int(room_id))
								t_day = DaysOfWeek.objects.filter(day_name = day_name)[0]
								t_div = Division.objects.filter(division = div)[0]
								if len(key_list) == 2: # it's a lec
									start_time, end_time = key_list
									t = TimeSlot.objects.filter(start_time = start_time, end_time = end_time)[0]
									#save lec here
									try:
										l = Lecture.objects.get(lec_day = t_day,lec_time = t,lec_div = t_div)
										if not (l.taken_by == u):
											l.taken_by = u
										if not (l.lname == s):
											l.lname = s
										if not (l.lec_in == r):
											l.lec_in = r
										l.save()
									except:
										l = Lecture(lname = s, lec_day = t_day, lec_time = t, lec_div = t_div, taken_by = u, lec_in = r)
										l.full_clean()
										l.save()

								elif len(key_list) == 3: # it's a lab
									start_time, end_time, batch_nm = key_list
									#save lab here
									t = TimeSlot.objects.filter(start_time = start_time, end_time = end_time)[0]
									batch = Batch.objects.filter(batch = batch_nm)[0]
									try:
										l = Lecture.objects.get(lec_day = t_day,lec_time = t,lec_div = t_div,lec_batch = batch)
										if not (l.taken_by == u):
											l.taken_by = u
										if not (l.lname == s):
											l.lname = s
										if not (l.lec_in == r):
											l.lec_in = r
										l.save()
									except:
										l = Lecture(lname = s, lec_day = t_day, lec_time = t, lec_div = t_div, taken_by = u, lec_in = r,lec_batch = batch)
										l.full_clean()
										l.save()
							else:
								t_day = DaysOfWeek.objects.filter(day_name = day_name)[0]
								t_div = Division.objects.filter(division = div)[0]
								t_year = Year.objects.filter(year = year)[0]

								if len(key_list) == 2: #delete a lec
									start_time, end_time = key_list
									t = TimeSlot.objects.filter(start_time = start_time, end_time = end_time)[0]
									try:
										l = Lecture.objects.get(lec_time = t, lec_day = t_day, lec_div = t_div, lname__year = t_year)
										l.delete()
									except: #throws Lectue.DoesNotExist
										pass
								elif len(key_list) == 3: #delete a lab
									start_time, end_time, batch_nm = key_list
									t = TimeSlot.objects.filter(start_time = start_time, end_time = end_time)[0]
									batch = Batch.objects.filter(batch = batch_nm)[0]
									try:
										l = Lecture.objects.get(lec_time = t, lec_day = t_day, lec_div = t_div, lname__year = t_year, lec_batch = batch)
										l.delete()
									except:
										pass
						else:
							if data_key == "yearopt":
								year = data_value[0]
							elif data_key == "termopt":
								term = data_value[0]
							elif data_key == "divopt":
								div = data_value[0]
				else:
					pass

		time_slots = TimeSlot.objects.annotate(
	    diff=ExpressionWrapper(F('end_time') - F('start_time'), output_field=DurationField())
		).filter(diff__lte=timedelta(hours = 2))

		pracs_ts = TimeSlot.objects.annotate(
	    diff=ExpressionWrapper(F('end_time') - F('start_time'), output_field=DurationField())
		).filter(diff=timedelta(hours = 2))

		subjects = Subject.objects.all();

		batches = Batch.objects.all();

		rooms = Room.objects.all();

		context_data = {
			"errors" : errors,
			"timeslots" : time_slots,
			"pracsts" : pracs_ts,
			"subjects" : subjects,
			"batches" : batches,
			"rooms" : rooms
		}
		return render(request,'assistant/updatett.html',context_data)
	html_error_data = {
		"error_code" : "401",
		"error_message" : "UNAUTHORIZED"
	}
	return render(request,"error.html",html_error_data)


def get_lec(request,year,division,timeslot,day,batch):
	if request.user.is_staff:
		year = Year.objects.get(year = year)
		# print(year)
		division = Division.objects.get(division = division)
		# print(division)
		timeslot = TimeSlot.objects.get(pk = int(timeslot))
		# print(timeslot)
		day = DaysOfWeek.objects.get(day_name = day)
		# print(day)

		try:
			lectures = list()
			if(batch == 'None'):
				lecture = Lecture.objects.get(lec_day = day,lec_time = timeslot,lec_div = division,lec_batch = None ,lname__year = year)
				lectures.append(lecture)
			else:
				batch = Batch.objects.get(pk = batch)
				# print(batch)
				lecture = Lecture.objects.get(lec_day = day,lec_time = timeslot,lec_div = division,lec_batch = batch,lname__year = year)
				lectures.append(lecture)

			lec_json = serializers.serialize("json",lectures)
			# print(lec_json)
			return HttpResponse(lec_json)
		except Exception as e:
			# print("Not found")
			response = JsonResponse({"error": "Not found"})
			# response.status_code = 403
			return response

	html_error_data = {
		"error_code" : "401",
		"error_message" : "UNAUTHORIZED"
	}
	return render(request,"error.html",html_error_data)


def get_preview_link(request):
	if request.user.is_staff:
		if request.POST:
			day_names = list(DaysOfWeek.objects.values_list('day_name', flat=True))[1:-1]
			form_data = OrderedDict()

			for name in day_names:
				stat_check = request.POST.get('form[' + name + ']', False)
				if stat_check:
					form_data[name] = QueryDict(request.POST.get('form[' + name + ']', False).encode('ASCII'))
				else:
					form_data[name] = QueryDict()

			# filepath = "static/upload/temp_tt.pdf"
			# og_filepath = "static/upload/tt.pdf"
			filepath = "static/upload/temp_timetable.xlsx"
			og_filepath = "static/upload/timetable.xlsx"
			if os.path.exists(filepath) and os.path.isfile(filepath):
				os.remove(filepath)

			thin_border = Border(left = Side(style='thin'),
								 right = Side(style='thin'),
								 top = Side(style='thin'),
								 bottom = Side(style='thin'))

			book = load_workbook(og_filepath)
			sheet = book.active
			column_letter = 'B'
			row_number = 13

			exclude_list = ['csrfmiddlewaretoken', 'termopt', 'yearopt', 'dayopt', 'divopt']
			subject_list = Subject.objects.values_list('id', 'sname')
			users_list = User.objects.values_list('id', 'username')
			room_list = Room.objects.values_list('id','room')
			year = ""
			div = ""
			term = ""
			ac_year = ""

			temp_count = 0
			for key, value in form_data.items():
				# increment x cord for diff days
				if temp_count == 0:
					column_letter = 'B'
				else:
					column_letter = chr(ord(column_letter) + 4)

				row_number = 13
				temp_count += 1
				lab_count = 0
				day_name = key
				qdict = value
				# print("=" * 30)
				# print(day_name, dict(qdict))
				# print("=" * 30)
				if len(qdict) > 0:
					for data_key, data_value in dict(qdict).items():

						if data_key not in exclude_list:
							key_list = list(map(lambda x: x.strip(), data_key.split('-')))

							lecture, room_id = data_value
							if lecture != '0' and room_id != 'None': # these are the values to be plotted on pdf
								sub_id, fac_id = lecture.split("-")
								sub_name = [sub[1] for sub in subject_list if sub[0] == int(sub_id)][0]
								fac_name = [fac[1] for fac in users_list if fac[0] == int(fac_id)][0]
								room_no = [room[1] for room in room_list if room[0] == int(room_id)][0]

								if len(key_list) == 2: # it's a lec
									start_time, end_time = key_list
									cell_range = f"{column_letter}{row_number}:{chr(ord(column_letter)+3)}{row_number}"
									sheet.merge_cells(cell_range)

									cell = sheet[f"{column_letter}{row_number}"]

									cell.value = sub_name + " " + room_no + " " + fac_name
									cell.alignment = Alignment(horizontal='center', vertical='center')
									cell.border = thin_border

									cell = sheet[f"{chr(ord(column_letter)+1)}{row_number}"]
									cell.border = thin_border
									cell = sheet[f"{chr(ord(column_letter)+2)}{row_number}"]
									cell.border = thin_border
									cell = sheet[f"{chr(ord(column_letter)+3)}{row_number}"]
									cell.border = thin_border
									#start
									if start_time == "10:30" and end_time == "11:30":
										row_number += 1
									elif start_time == "11:30" and end_time == "12:30":
										row_number += 2
									#mid
									elif start_time == "13:15" and end_time == "14:15":
										row_number += 1
									elif start_time == "14:15" and end_time == "15:15":
										row_number += 1
									#end
									elif start_time == "15:15" and end_time == "16:15":
										row_number += 1
									elif start_time == "16:15" and end_time == "17:15":
										row_number += 2

								elif len(key_list) == 3: # it's a lab
									start_time, end_time, batch = key_list

									cell_range = f"{chr(ord(column_letter) + lab_count)}{row_number}:{chr(ord(column_letter) + lab_count)}{row_number+1}"
									sheet.merge_cells(cell_range)

									cell = sheet[f"{chr(ord(column_letter) + lab_count)}{row_number}"]
									cell.value = sub_name + " " + room_no + " " + fac_name + " " + batch
									cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

									cell.border = thin_border

									cell = sheet[f"{chr(ord(column_letter) + lab_count)}{row_number+1}"]
									cell.border = thin_border

									lab_count += 1
									#start
									if start_time == "10:30" and end_time == "12:30":
										if lab_count % 4 == 0:
											row_number += 3
									#mid
									elif start_time == "13:15" and end_time == "15:15":
										if lab_count % 4 == 0:
											row_number += 2
									#end
									elif start_time == "15:15" and end_time == "17:15":
										if lab_count % 4 == 0:
											row_number += 2

									if lab_count == 4:
										lab_count = 0
									div = batch[0]
						else:
							if data_key == "yearopt":
								year = data_value[0]
							elif data_key == "termopt":
								term = data_value[0]
				else:
					pass
					# print("+" * 30)
					# print(day_name)
					# print("+" * 30)

			#date on top
			cell = sheet["A1"]
			cell.value = datetime.today().strftime(r"%d/%m/%Y")
			cell.alignment = Alignment(horizontal='center', vertical='center')

			#class
			cell = sheet["B9"]
			cell.value = f"{year.upper()} B.Tech IT {div.upper()}"
			cell.alignment = Alignment(horizontal='left', vertical='center')

			#semester
			cell = sheet["U9"]
			cell.value = term.upper()
			cell.alignment = Alignment(horizontal='left', vertical='center')

			curr_year = int(datetime.now().year)
			if term.upper() == "ODD":
				ac_year = str(curr_year) + '_' + str(curr_year+1)
			elif term.upper() == "EVEN":
				ac_year = str(curr_year-1) + '_' + str(curr_year)

			#academic year
			cell = sheet["P9"]
			cell.value = ac_year
			cell.alignment = Alignment(horizontal='left', vertical='center')

			book.save(filepath)
			api = cloudconvert.Api(settings.CLOUDCONVERT_API_KEY)

			process = api.convert({
				"inputformat": "xlsx",
				"outputformat": "pdf",
				"input": "upload",
				"file": open(filepath, 'rb')
			})
			process.wait()
			filepath = "static/upload/temp_timetable.pdf"
			process.download(filepath)

			# print("=" * 30)
			# print(form_data)
			# print("=" * 30)

			return HttpResponse(filepath)

	html_error_data = {
		"error_code" : "401",
		"error_message" : "UNAUTHORIZED"
	}
	return render(request,"error.html",html_error_data)
